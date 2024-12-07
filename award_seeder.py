import openpyxl
from sqlalchemy import create_engine, Column, Integer, String, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship

# Define the models
Base = declarative_base()

class Assistant(Base):
    __tablename__ = 'assistants'
    id = Column(Integer, primary_key=True, autoincrement=True)
    initial = Column(String(10), unique=True, nullable=False)
    full_name = Column(String)

class Award(Base):
    __tablename__ = 'awards'
    id = Column(Integer, primary_key=True, autoincrement=True)
    award_title = Column(String(100), unique=True, nullable=False)
    award_description = Column(String)

class Period(Base):
    __tablename__ = 'periods'
    id = Column(Integer, primary_key=True, autoincrement=True)
    period_title = Column(String(50), unique=True, nullable=False)

class AssistantAward(Base):
    __tablename__ = 'assistant_awards'
    id = Column(Integer, primary_key=True, autoincrement=True)
    assistant_id = Column(Integer, ForeignKey('assistants.id'), nullable=False)
    award_id = Column(Integer, ForeignKey('awards.id'), nullable=False)
    period_id = Column(Integer, ForeignKey('periods.id'), nullable=False)
    award_image = Column(String)

DATABASE_URL = "postgresql://abel:hehe@localhost:5432/slc_website"
engine = create_engine(DATABASE_URL)
Session = sessionmaker(bind=engine)
session = Session()

Base.metadata.create_all(engine)

def normalize_semester(semester):
    """Normalize the semester format for comparison."""
    parts = semester.split()
    if len(parts) == 2:  # Handles "Spring 2023" cases
        return f"{parts[0]} Semester {parts[1]}"
    elif len(parts) == 3 and parts[0].lower() in {"odd", "even"}:  # Handles "odd 2022/2023"
        return f"{parts[0].capitalize()} Semester {parts[1]}/{parts[2]}"
    return semester  # Return as is if it doesn't match the expected format


def insert_assistant_awards_from_xlsx(file_path):
    try:
        # Open the Excel workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Get the first sheet

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            # Debug: Print the row to verify its structure
            print(f"Processing row: {row}")

            # Check for invalid rows (skip rows with missing Initial or required columns)
            if row[3] is None or row[4] is None or row[5] is None:
                print(f"Skipping row with missing or invalid data: {row}")
                continue

            # Extract data
            name = str(row[3]).strip() if row[3] else None
            award_name = str(row[4]).strip()
            semester = normalize_semester(str(row[5]).strip())

            # Validate Initial
            if not name:
                print(f"Skipping row with missing or invalid name: {row}")
                continue

            # Find the assistant by initial
            assistant = session.query(Assistant).filter_by(full_name=name).first()
            if not assistant:
                print(f"Assistant with name '{name}' not found. Skipping row.")
                continue

            # Find the award by award name
            award = session.query(Award).filter_by(award_title=award_name).first()
            if not award:
                print(f"Award with title '{award_name}' not found. Skipping row.")
                continue

            # Find the period by normalized semester name
            period = session.query(Period).filter_by(period_title=semester).first()
            if not period:
                print(f"Period with name '{semester}' not found. Skipping row.")
                continue

            # Check if the assistant-award record already exists
            existing_entry = session.query(AssistantAward).filter_by(
                assistant_id=assistant.id, award_id=award.id, period_id=period.id
            ).first()

            if existing_entry:
                print(f"AssistantAward for Assistant '{name}', Award '{award_name}', and Period '{semester}' already exists. Skipping.")
                continue

            # Insert the assistant-award relationship
            new_assistant_award = AssistantAward(
                assistant_id=assistant.id,
                award_id=award.id,
                period_id=period.id,
                award_image="",  # Add image path if needed
            )
            session.add(new_assistant_award)
            print(f"Inserted AssistantAward for Assistant '{name}', Award '{award_name}', and Period '{semester}'.")

        # Commit the transaction
        session.commit()
        print("AssistantAwards insertion completed.")

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        session.rollback()

def insert_awards_from_xlsx(file_path):
    try:
        # Open the Excel workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Get the first sheet

        award_set = set()  # To track unique awards
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            if len(row) < 5:
                print(f"Skipping invalid row: {row}")
                continue

            # Extract the award name
            award_name = str(row[4]).strip()  # Column 4: Award Name
            if award_name == "None" or award_name == "Award Name":
                continue
            if award_name not in award_set:
                award_set.add(award_name)

                # Check if award already exists in the database
                existing_award = session.query(Award).filter_by(award_title=award_name).first()
                if not existing_award:
                    new_award = Award(award_title=award_name, award_description="")
                    session.add(new_award)
                    print(f"Inserted award: {award_name}")
                else:
                    print(f"Award already exists: {award_name}")

        # Commit the transaction
        session.commit()
        print("Awards insertion completed.")

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        session.rollback()


if __name__ == "__main__":
    xlsx_file_path = "DataAwardRecap.xlsx"  # Replace with your .xlsx file path
    insert_awards_from_xlsx(xlsx_file_path)
    insert_assistant_awards_from_xlsx(xlsx_file_path)
